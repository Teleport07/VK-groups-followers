import time
import openpyxl
import requests

#подготавливаем excel файл, на первом листе которого будет список групп, id участников которых нужно получить
wb = openpyxl.load_workbook('путь к excel файлу')
ws_begin = wb['Лист1']
ws = wb['Лист2']

#в part2 вводим личный vk токен, полученный после инициализации приложения через в vk api
part2 = "vk токен"
paginator = "0"
count_row = 0

#в цикле for пробегаем по строкам первого листа файла excel с наименованиями целевых групп
for url in range(1, 2):
    count_row += 1
    ws.cell(row=1, column=count_row).value = ws_begin.cell(row=url, column=1).value
    count = 2
    part1 = "https://api.vk.com/method/groups.getMembers?group_id="+ws_begin.cell(row = url, column = 1).value.split('.com/')[1]+"&count=1000&start_from="
    while True:
        r = requests.get(part1+paginator+part2)
        r1 = r.text.split(':[')[1].split(',')
        print(r.content)
        for i in r1:
            if count % 1000000 == 0:
                count = 1
                count_row += 1
            if i.endswith(']'):
                ws.cell(row=count, column=count_row).value = i.split(']')[0]
                count += 1
                break
            ws.cell(row = count, column = count_row).value = i
            count += 1
        time.sleep(0.5)
        paginator = r.text.split('from":"')[1].split('"')[0]
        if r.text.endswith('"next_from":""}}'):
            break

#как итог получаем на втором листе столбцы со спаршенными id участников групп.
#Каждый столбец соответствует одной группе

wb.save('путь к excel файлу')